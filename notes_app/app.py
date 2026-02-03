"""Retro Notes Manager - Flask app using Excel as a database."""
from __future__ import annotations

import base64
import os
import re
from datetime import datetime
from io import BytesIO
from typing import List, Dict

from flask import Flask, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.fernet import Fernet

APP_ROOT = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_ROOT, "notes_database.xlsx")

app = Flask(__name__)


HEADERS = ["ID", "Title", "Content", "DateCreated", "LastModified"]


def ensure_database() -> None:
    """Create the Excel database with headers if it does not exist."""
    if os.path.exists(DB_PATH):
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet.append(HEADERS)
    workbook.save(DB_PATH)


def get_sheet(workbook) -> Worksheet:
    """Return the primary sheet that stores notes."""
    return workbook.active


def load_notes() -> List[Dict[str, str]]:
    """Load all notes from Excel into a list of dictionaries."""
    ensure_database()
    workbook = load_workbook(DB_PATH)
    sheet = get_sheet(workbook)
    notes: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        notes.append(
            {
                "id": str(row[0]),
                "title": row[1] or "",
                "content": row[2] or "",
                "date_created": row[3] or "",
                "last_modified": row[4] or "",
            }
        )
    workbook.close()
    return notes


def get_next_id(notes: List[Dict[str, str]]) -> int:
    """Return the next available integer ID."""
    if not notes:
        return 1
    return max(int(note["id"]) for note in notes) + 1


def save_note(note_id: str | None, title: str, content: str) -> str:
    """Save a new note or update an existing one."""
    ensure_database()
    workbook = load_workbook(DB_PATH)
    sheet = get_sheet(workbook)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if note_id:
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == str(note_id):
                row[1].value = title
                row[2].value = content
                row[4].value = now
                workbook.save(DB_PATH)
                workbook.close()
                return str(note_id)

    notes = load_notes()
    new_id = get_next_id(notes)
    sheet.append([new_id, title, content, now, now])
    workbook.save(DB_PATH)
    workbook.close()
    return str(new_id)


def delete_note(note_id: str) -> None:
    """Delete a note by ID."""
    ensure_database()
    workbook = load_workbook(DB_PATH)
    sheet = get_sheet(workbook)
    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == str(note_id):
            sheet.delete_rows(row[0].row, 1)
            break
    workbook.save(DB_PATH)
    workbook.close()


def find_note(note_id: str) -> Dict[str, str] | None:
    """Return a single note by ID."""
    notes = load_notes()
    for note in notes:
        if note["id"] == str(note_id):
            return note
    return None


def sanitize_filename(title: str) -> str:
    """Sanitize a note title for file output."""
    safe = re.sub(r"[^a-zA-Z0-9_-]+", "_", title.strip())
    return safe or "untitled"


def derive_key(password: str, salt: bytes) -> bytes:
    """Derive a Fernet key from a password and salt."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=390000,
    )
    return base64.urlsafe_b64encode(kdf.derive(password.encode("utf-8")))


def encrypt_note(note: Dict[str, str], password: str) -> bytes:
    """Encrypt a note with a password and return file bytes."""
    salt = os.urandom(16)
    key = derive_key(password, salt)
    fernet = Fernet(key)
    plaintext = (
        f"Title: {note['title']}\n"
        f"Created: {note['date_created']}\n"
        f"Last Modified: {note['last_modified']}\n\n"
        f"{note['content']}"
    )
    token = fernet.encrypt(plaintext.encode("utf-8"))
    payload = base64.b64encode(salt) + b"\n" + token
    return payload


@app.route("/")
def index():
    notes = load_notes()
    return render_template("index.html", notes=notes)


@app.route("/note/<note_id>")
def view_note(note_id: str):
    note = find_note(note_id)
    if not note:
        return redirect(url_for("index"))
    return render_template("editor.html", note=note)


@app.route("/save", methods=["POST"])
def save():
    note_id = request.form.get("note_id")
    title = request.form.get("title", "Untitled").strip() or "Untitled"
    content = request.form.get("content", "").strip()
    saved_id = save_note(note_id, title, content)
    return redirect(url_for("view_note", note_id=saved_id))


@app.route("/delete/<note_id>", methods=["POST"])
def delete(note_id: str):
    delete_note(note_id)
    return redirect(url_for("index"))


@app.route("/search")
def search():
    query = request.args.get("q", "").strip().lower()
    notes = load_notes()
    if not query:
        return jsonify(notes)
    matches = [
        note
        for note in notes
        if query in note["title"].lower() or query in note["content"].lower()
    ]
    return jsonify(matches)


@app.route("/export/<note_id>", methods=["POST"])
def export(note_id: str):
    password = request.form.get("password", "")
    if not password:
        return redirect(url_for("view_note", note_id=note_id))
    note = find_note(note_id)
    if not note:
        return redirect(url_for("index"))
    payload = encrypt_note(note, password)
    filename = f"note_{sanitize_filename(note['title'])}.secure"
    return send_file(
        BytesIO(payload),
        as_attachment=True,
        download_name=filename,
        mimetype="application/octet-stream",
    )


if __name__ == "__main__":
    ensure_database()
    app.run(host="0.0.0.0", port=5000, debug=True)
